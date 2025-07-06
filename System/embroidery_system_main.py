#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
刺绣工资管理系统 - 主程序
作者：Claude
功能：完整的刺绣工资管理系统，包含员工管理、花型管理、工资记录、统计报表等功能
"""

import webview
import sqlite3
import json
import base64
import os
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO

class EmbroiderySystem:
    """刺绣工资管理系统核心类"""
    
    def __init__(self):
        """初始化系统"""
        self.db_path = 'embroidery_system.db'
        self.init_database()
        self.init_default_patterns()
    
    def init_database(self):
        """初始化数据库表结构"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 创建员工表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                status TEXT DEFAULT 'active',
                create_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 创建花型表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS patterns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                price REAL NOT NULL,
                create_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 创建工资记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                pattern_id INTEGER NOT NULL,
                record_date DATE NOT NULL,
                count INTEGER NOT NULL,
                total REAL NOT NULL,
                special TEXT DEFAULT '',
                note TEXT DEFAULT '',
                create_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees(id),
                FOREIGN KEY (pattern_id) REFERENCES patterns(id)
            )
        ''')
        
        # 创建设置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def init_default_patterns(self):
        """初始化默认花型数据"""
        default_patterns = [
            ('盖布', 5.5), ('浮花', 5.5), ('棉盖', 5.5), ('棉', 5.0),
            ('藏网', 5.5), ('雕孔', 6.5), ('棉网', 5.5), ('水溶', 4.4),
            ('牛奶丝打孔', 6.5), ('网布', 5.5), ('玻璃纱', 5.2), ('盖网', 5.0)
        ]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        for name, price in default_patterns:
            cursor.execute('''
                INSERT OR IGNORE INTO patterns (name, price)
                VALUES (?, ?)
            ''', (name, price))
        
        conn.commit()
        conn.close()
    
    # ==================== 员工管理 API ====================
    
    def get_employees(self):
        """获取所有员工列表"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, name, status, create_date 
            FROM employees 
            ORDER BY create_date DESC
        ''')
        
        employees = []
        for row in cursor.fetchall():
            employees.append({
                'id': row[0],
                'name': row[1],
                'status': row[2],
                'create_date': row[3]
            })
        
        conn.close()
        return {'success': True, 'data': employees}
    
    def add_employee(self, name):
        """添加员工"""
        if not name or name.strip() == '':
            return {'success': False, 'message': '员工姓名不能为空'}
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('INSERT INTO employees (name) VALUES (?)', (name.strip(),))
            conn.commit()
            return {'success': True, 'message': '员工添加成功'}
        except sqlite3.IntegrityError:
            return {'success': False, 'message': '员工姓名已存在'}
        finally:
            conn.close()
    
    def update_employee(self, emp_id, name, status):
        """更新员工信息"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                UPDATE employees 
                SET name = ?, status = ? 
                WHERE id = ?
            ''', (name, status, emp_id))
            
            if cursor.rowcount > 0:
                conn.commit()
                return {'success': True, 'message': '员工信息更新成功'}
            else:
                return {'success': False, 'message': '员工不存在'}
        except sqlite3.IntegrityError:
            return {'success': False, 'message': '员工姓名已存在'}
        finally:
            conn.close()
    
    def delete_employee(self, emp_id):
        """删除员工"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 检查是否有相关工资记录
        cursor.execute('SELECT COUNT(*) FROM records WHERE employee_id = ?', (emp_id,))
        record_count = cursor.fetchone()[0]
        
        if record_count > 0:
            conn.close()
            return {'success': False, 'message': '该员工有工资记录，无法删除'}
        
        cursor.execute('DELETE FROM employees WHERE id = ?', (emp_id,))
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            return {'success': True, 'message': '员工删除成功'}
        else:
            conn.close()
            return {'success': False, 'message': '员工不存在'}
    
    # ==================== 花型管理 API ====================
    
    def get_patterns(self):
        """获取所有花型列表"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, name, price, create_date 
            FROM patterns 
            ORDER BY create_date DESC
        ''')
        
        patterns = []
        for row in cursor.fetchall():
            patterns.append({
                'id': row[0],
                'name': row[1],
                'price': row[2],
                'create_date': row[3]
            })
        
        conn.close()
        return {'success': True, 'data': patterns}
    
    def add_pattern(self, name, price):
        """添加花型"""
        if not name or name.strip() == '':
            return {'success': False, 'message': '花型名称不能为空'}
        
        try:
            price = float(price)
            if price <= 0:
                return {'success': False, 'message': '单价必须大于0'}
        except ValueError:
            return {'success': False, 'message': '单价格式错误'}
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('INSERT INTO patterns (name, price) VALUES (?, ?)', 
                         (name.strip(), price))
            conn.commit()
            return {'success': True, 'message': '花型添加成功'}
        except sqlite3.IntegrityError:
            return {'success': False, 'message': '花型名称已存在'}
        finally:
            conn.close()
    
    def update_pattern(self, pattern_id, name, price):
        """更新花型信息"""
        try:
            price = float(price)
            if price <= 0:
                return {'success': False, 'message': '单价必须大于0'}
        except ValueError:
            return {'success': False, 'message': '单价格式错误'}
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                UPDATE patterns 
                SET name = ?, price = ? 
                WHERE id = ?
            ''', (name, price, pattern_id))
            
            if cursor.rowcount > 0:
                conn.commit()
                return {'success': True, 'message': '花型信息更新成功'}
            else:
                return {'success': False, 'message': '花型不存在'}
        except sqlite3.IntegrityError:
            return {'success': False, 'message': '花型名称已存在'}
        finally:
            conn.close()
    
    def delete_pattern(self, pattern_id):
        """删除花型"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 检查是否有相关工资记录
        cursor.execute('SELECT COUNT(*) FROM records WHERE pattern_id = ?', (pattern_id,))
        record_count = cursor.fetchone()[0]
        
        if record_count > 0:
            conn.close()
            return {'success': False, 'message': '该花型有工资记录，无法删除'}
        
        cursor.execute('DELETE FROM patterns WHERE id = ?', (pattern_id,))
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            return {'success': True, 'message': '花型删除成功'}
        else:
            conn.close()
            return {'success': False, 'message': '花型不存在'}
    
    # ==================== 工资记录管理 API ====================
    
    def get_records(self, employee_id=None, pattern_id=None, start_date=None, end_date=None):
        """获取工资记录列表"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 构建查询条件
        conditions = []
        params = []
        
        if employee_id:
            conditions.append('r.employee_id = ?')
            params.append(employee_id)
        
        if pattern_id:
            conditions.append('r.pattern_id = ?')
            params.append(pattern_id)
        
        if start_date:
            conditions.append('r.record_date >= ?')
            params.append(start_date)
        
        if end_date:
            conditions.append('r.record_date <= ?')
            params.append(end_date)
        
        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        
        query = f'''
            SELECT r.id, r.employee_id, r.pattern_id, r.record_date, 
                   r.count, r.total, r.special, r.note, r.create_time,
                   e.name as employee_name, p.name as pattern_name, p.price
            FROM records r
            JOIN employees e ON r.employee_id = e.id
            JOIN patterns p ON r.pattern_id = p.id
            WHERE {where_clause}
            ORDER BY r.record_date DESC, r.create_time DESC
        '''
        
        cursor.execute(query, params)
        
        records = []
        for row in cursor.fetchall():
            records.append({
                'id': row[0],
                'employee_id': row[1],
                'pattern_id': row[2],
                'record_date': row[3],
                'count': row[4],
                'total': row[5],
                'special': row[6],
                'note': row[7],
                'create_time': row[8],
                'employee_name': row[9],
                'pattern_name': row[10],
                'price': row[11]
            })
        
        conn.close()
        return {'success': True, 'data': records}
    
    def add_record(self, employee_id, pattern_id, record_date, count, special='', note=''):
        """添加工资记录"""
        try:
            count = int(count)
            if count <= 0:
                return {'success': False, 'message': '针数必须大于0'}
        except ValueError:
            return {'success': False, 'message': '针数格式错误'}
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 获取花型单价
        cursor.execute('SELECT price FROM patterns WHERE id = ?', (pattern_id,))
        pattern_result = cursor.fetchone()
        if not pattern_result:
            conn.close()
            return {'success': False, 'message': '花型不存在'}
        
        price = pattern_result[0]
        total = count * price
        
        try:
            cursor.execute('''
                INSERT INTO records (employee_id, pattern_id, record_date, count, total, special, note)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (employee_id, pattern_id, record_date, count, total, special, note))
            
            conn.commit()
            return {'success': True, 'message': '工资记录添加成功'}
        except Exception as e:
            return {'success': False, 'message': f'添加失败: {str(e)}'}
        finally:
            conn.close()
    
    def update_record(self, record_id, employee_id, pattern_id, record_date, count, special='', note=''):
        """更新工资记录"""
        try:
            count = int(count)
            if count <= 0:
                return {'success': False, 'message': '针数必须大于0'}
        except ValueError:
            return {'success': False, 'message': '针数格式错误'}
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 获取花型单价
        cursor.execute('SELECT price FROM patterns WHERE id = ?', (pattern_id,))
        pattern_result = cursor.fetchone()
        if not pattern_result:
            conn.close()
            return {'success': False, 'message': '花型不存在'}
        
        price = pattern_result[0]
        total = count * price
        
        try:
            cursor.execute('''
                UPDATE records 
                SET employee_id = ?, pattern_id = ?, record_date = ?, 
                    count = ?, total = ?, special = ?, note = ?
                WHERE id = ?
            ''', (employee_id, pattern_id, record_date, count, total, special, note, record_id))
            
            if cursor.rowcount > 0:
                conn.commit()
                return {'success': True, 'message': '工资记录更新成功'}
            else:
                return {'success': False, 'message': '记录不存在'}
        except Exception as e:
            return {'success': False, 'message': f'更新失败: {str(e)}'}
        finally:
            conn.close()
    
    def delete_record(self, record_id):
        """删除工资记录"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM records WHERE id = ?', (record_id,))
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            return {'success': True, 'message': '工资记录删除成功'}
        else:
            conn.close()
            return {'success': False, 'message': '记录不存在'}
    
    # ==================== 统计报表 API ====================
    
    def get_daily_summary(self, date):
        """获取日工资汇总"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT e.name as employee_name, p.name as pattern_name, 
                   SUM(r.count) as total_count, SUM(r.total) as total_wage
            FROM records r
            JOIN employees e ON r.employee_id = e.id
            JOIN patterns p ON r.pattern_id = p.id
            WHERE r.record_date = ?
            GROUP BY e.name, p.name
            ORDER BY e.name, p.name
        ''', (date,))
        
        summary = []
        for row in cursor.fetchall():
            summary.append({
                'employee_name': row[0],
                'pattern_name': row[1],
                'total_count': row[2],
                'total_wage': row[3]
            })
        
        conn.close()
        return {'success': True, 'data': summary}
    
    def get_monthly_summary(self, year, month):
        """获取月工资汇总"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 获取指定月份的数据
        start_date = f'{year}-{month:02d}-01'
        if month == 12:
            end_date = f'{year+1}-01-01'
        else:
            end_date = f'{year}-{month+1:02d}-01'
        
        cursor.execute('''
            SELECT e.name as employee_name, p.name as pattern_name, 
                   SUM(r.count) as total_count, SUM(r.total) as total_wage
            FROM records r
            JOIN employees e ON r.employee_id = e.id
            JOIN patterns p ON r.pattern_id = p.id
            WHERE r.record_date >= ? AND r.record_date < ?
            GROUP BY e.name, p.name
            ORDER BY e.name, p.name
        ''', (start_date, end_date))
        
        # 构建汇总数据结构
        summary_data = {}
        patterns = set()
        
        for row in cursor.fetchall():
            employee_name = row[0]
            pattern_name = row[1]
            total_count = row[2]
            total_wage = row[3]
            
            if employee_name not in summary_data:
                summary_data[employee_name] = {}
            
            summary_data[employee_name][pattern_name] = {
                'count': total_count,
                'wage': total_wage
            }
            patterns.add(pattern_name)
        
        conn.close()
        return {
            'success': True, 
            'data': {
                'summary': summary_data,
                'patterns': sorted(list(patterns))
            }
        }
    
    def export_monthly_excel(self, year, month):
        """导出月工资汇总Excel"""
        try:
            # 获取月汇总数据
            result = self.get_monthly_summary(year, month)
            if not result['success']:
                return {'success': False, 'message': '获取数据失败'}
            
            summary_data = result['data']['summary']
            patterns = result['data']['patterns']
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = f'{year}年{month:02d}月工资汇总'
            
            # 设置样式
            header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            center_alignment = Alignment(horizontal='center', vertical='center')
            total_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # 设置标题行
            headers = ['员工姓名'] + patterns + ['合计']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.border = border
                cell.alignment = center_alignment
            
            # 填充数据
            row_num = 2
            total_by_pattern = {pattern: {'count': 0, 'wage': 0} for pattern in patterns}
            
            for employee_name, employee_data in summary_data.items():
                ws.cell(row=row_num, column=1, value=employee_name)
                
                row_total_count = 0
                row_total_wage = 0
                
                for col_num, pattern in enumerate(patterns, 2):
                    if pattern in employee_data:
                        count = employee_data[pattern]['count']
                        wage = employee_data[pattern]['wage']
                        cell_value = f'{count}/{wage:.2f}'
                        
                        row_total_count += count
                        row_total_wage += wage
                        total_by_pattern[pattern]['count'] += count
                        total_by_pattern[pattern]['wage'] += wage
                    else:
                        cell_value = '0/0.00'
                    
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.border = border
                    cell.alignment = center_alignment
                
                # 行合计
                total_cell = ws.cell(row=row_num, column=len(patterns) + 2, 
                                   value=f'{row_total_count}/{row_total_wage:.2f}')
                total_cell.border = border
                total_cell.alignment = center_alignment
                
                # 设置第一列边框和对齐
                first_cell = ws.cell(row=row_num, column=1)
                first_cell.border = border
                first_cell.alignment = center_alignment
                
                row_num += 1
            
            # 添加合计行
            ws.cell(row=row_num, column=1, value='合计')
            total_count = 0
            total_wage = 0
            
            for col_num, pattern in enumerate(patterns, 2):
                pattern_total = total_by_pattern[pattern]
                cell_value = f'{pattern_total["count"]}/{pattern_total["wage"]:.2f}'
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
                cell.alignment = center_alignment
                cell.fill = total_fill
                cell.font = header_font
                
                total_count += pattern_total["count"]
                total_wage += pattern_total["wage"]
            
            # 总合计
            total_cell = ws.cell(row=row_num, column=len(patterns) + 2, 
                               value=f'{total_count}/{total_wage:.2f}')
            total_cell.border = border
            total_cell.alignment = center_alignment
            total_cell.fill = total_fill
            total_cell.font = header_font
            
            # 设置合计行第一列样式
            first_cell = ws.cell(row=row_num, column=1)
            first_cell.border = border
            first_cell.alignment = center_alignment
            first_cell.fill = total_fill
            first_cell.font = header_font
            
            # 调整列宽
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15
            
            # 保存到内存
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # 转换为base64
            excel_base64 = base64.b64encode(excel_buffer.read()).decode()
            filename = f'工资汇总_{year}年{month:02d}月.xlsx'
            
            return {
                'success': True,
                'data': {
                    'filename': filename,
                    'content': excel_base64
                }
            }
        
        except Exception as e:
            return {'success': False, 'message': f'导出失败: {str(e)}'}
    
    def import_excel_data(self, file_content):
        """导入Excel数据"""
        try:
            # 解码base64文件内容
            file_data = base64.b64decode(file_content)
            
            # 使用pandas读取Excel
            df = pd.read_excel(BytesIO(file_data))
            
            # 验证必要的列
            required_columns = ['员工姓名', '花型', '针数', '记录日期']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return {'success': False, 'message': f'缺少必要的列: {", ".join(missing_columns)}'}
            
            # 开始导入数据
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            success_count = 0
            error_records = []
            
            for index, row in df.iterrows():
                try:
                    employee_name = str(row['员工姓名']).strip()
                    pattern_name = str(row['花型']).strip()
                    count = int(row['针数'])
                    record_date = pd.to_datetime(row['记录日期']).strftime('%Y-%m-%d')
                    special = str(row.get('特殊说明', '')).strip()
                    note = str(row.get('备注', '')).strip()
                    
                    # 查找员工ID
                    cursor.execute('SELECT id FROM employees WHERE name = ?', (employee_name,))
                    employee_result = cursor.fetchone()
                    if not employee_result:
                        # 自动创建员工
                        cursor.execute('INSERT INTO employees (name) VALUES (?)', (employee_name,))
                        employee_id = cursor.lastrowid
                    else:
                        employee_id = employee_result[0]
                    
                    # 查找花型ID
                    cursor.execute('SELECT id, price FROM patterns WHERE name = ?', (pattern_name,))
                    pattern_result = cursor.fetchone()
                    if not pattern_result:
                        error_records.append(f'第{index+2}行: 花型"{pattern_name}"不存在')
                        continue
                    
                    pattern_id = pattern_result[0]
                    price = pattern_result[1]
                    total = count * price
                    
                    # 插入记录
                    cursor.execute('''
                        INSERT INTO records (employee_id, pattern_id, record_date, count, total, special, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (employee_id, pattern_id, record_date, count, total, special, note))
                    
                    success_count += 1
                    
                except Exception as e:
                    error_records.append(f'第{index+2}行: {str(e)}')
                    continue
            
            conn.commit()
            conn.close()
            
            message = f'成功导入 {success_count} 条记录'
            if error_records:
                message += f'，{len(error_records)} 条记录失败'
            
            return {
                'success': True,
                'message': message,
                'error_records': error_records
            }
            
        except Exception as e:
            return {'success': False, 'message': f'导入失败: {str(e)}'}
    
    # ==================== 设置管理 API ====================
    
    def get_setting(self, key):
        """获取设置"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('SELECT value FROM settings WHERE key = ?', (key,))
        result = cursor.fetchone()
        
        conn.close()
        return result[0] if result else None
    
    def set_setting(self, key, value):
        """设置配置"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT OR REPLACE INTO settings (key, value)
            VALUES (?, ?)
        ''', (key, value))
        
        conn.commit()
        conn.close()
        return {'success': True, 'message': '设置保存成功'}


# 创建系统实例
system = EmbroiderySystem()


def main():
    """主函数"""
    # 获取资源路径
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller打包后的路径
        bundle_dir = sys._MEIPASS
    else:
        # 开发环境路径
        bundle_dir = os.path.dirname(os.path.abspath(__file__))
    
    html_path = os.path.join(bundle_dir, 'index.html')
    
    # 创建webview窗口
    webview.create_window(
        title='刺绣工资管理系统',
        url=html_path,
        js_api=system,
        width=1200,
        height=800,
        min_size=(1000, 600),
        resizable=True
    )
    
    # 启动应用
    webview.start(debug=False)


if __name__ == '__main__':
    main()
