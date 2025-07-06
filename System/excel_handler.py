import io
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import calendar
from typing import Dict, List, Any

class ExcelHandler:
    def __init__(self, db):
        """
        初始化Excel处理器
        
        Args:
            db: 数据库实例
        """
        self.db = db
    
    def import_records(self, file_data: bytes) -> Dict[str, Any]:
        """
        从Excel文件导入工资记录
        
        Args:
            file_data: Excel文件二进制数据
            
        Returns:
            导入结果
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(io.BytesIO(file_data))
            
            # 验证必要的列
            required_columns = ['员工姓名', '花型名称', '记录日期', '针数']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return {
                    'success': False,
                    'message': f'Excel文件缺少必要列: {", ".join(missing_columns)}'
                }
            
            # 获取员工和花型映射
            employees = self.db.get_employees()
            patterns = self.db.get_patterns()
            
            employee_map = {emp['name']: emp['id'] for emp in employees}
            pattern_map = {pat['name']: pat['id'] for pat in patterns}
            
            success_count = 0
            error_count = 0
            error_messages = []
            
            # 处理每一行数据
            for index, row in df.iterrows():
                try:
                    employee_name = str(row['员工姓名']).strip()
                    pattern_name = str(row['花型名称']).strip()
                    record_date = str(row['记录日期'])
                    count = int(row['针数'])
                    
                    # 验证数据
                    if employee_name not in employee_map:
                        error_messages.append(f'第{index+2}行: 员工"{employee_name}"不存在')
                        error_count += 1
                        continue
                    
                    if pattern_name not in pattern_map:
                        error_messages.append(f'第{index+2}行: 花型"{pattern_name}"不存在')
                        error_count += 1
                        continue
                    
                    if count <= 0:
                        error_messages.append(f'第{index+2}行: 针数必须大于0')
                        error_count += 1
                        continue
                    
                    # 处理日期格式
                    try:
                        if isinstance(row['记录日期'], pd.Timestamp):
                            record_date = row['记录日期'].strftime('%Y-%m-%d')
                        else:
                            # 尝试解析日期字符串
                            parsed_date = pd.to_datetime(record_date)
                            record_date = parsed_date.strftime('%Y-%m-%d')
                    except:
                        error_messages.append(f'第{index+2}行: 日期格式错误')
                        error_count += 1
                        continue
                    
                    # 获取其他字段
                    special = str(row.get('特殊标记', '')).strip()
                    note = str(row.get('备注', '')).strip()
                    
                    # 添加记录
                    result = self.db.add_record(
                        employee_id=employee_map[employee_name],
                        pattern_id=pattern_map[pattern_name],
                        record_date=record_date,
                        count=count,
                        special=special,
                        note=note
                    )
                    
                    if result['success']:
                        success_count += 1
                    else:
                        error_messages.append(f'第{index+2}行: {result["message"]}')
                        error_count += 1
                
                except Exception as e:
                    error_messages.append(f'第{index+2}行: 处理失败 - {str(e)}')
                    error_count += 1
            
            return {
                'success': True,
                'message': f'导入完成: 成功{success_count}条，失败{error_count}条',
                'data': {
                    'success_count': success_count,
                    'error_count': error_count,
                    'error_messages': error_messages[:10]  # 只返回前10条错误信息
                }
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'导入失败: {str(e)}'
            }
    
    def export_monthly_summary(self, year: int, month: int) -> Dict[str, Any]:
        """
        导出月度工资汇总Excel
        
        Args:
            year: 年份
            month: 月份
            
        Returns:
            导出结果
        """
        try:
            # 获取月度汇总数据
            summary_data = self.db.get_monthly_summary(year, month)
            
            if not summary_data['data']:
                return {
                    'success': False,
                    'message': '该月份没有工资记录'
                }
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = f"{year}年{month}月工资汇总"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 准备数据
            data = summary_data['data']
            patterns = summary_data['patterns']
            employees = sorted(data.keys())
            
            # 创建表头
            headers = ['员工姓名'] + patterns + ['合计']
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # 写入数据行
            row_num = 2
            pattern_totals = {pattern: {'count': 0, 'amount': 0} for pattern in patterns}
            
            for employee in employees:
                ws.cell(row=row_num, column=1, value=employee).border = border
                
                employee_total = 0
                
                for col, pattern in enumerate(patterns, 2):
                    if pattern in data[employee]:
                        count = data[employee][pattern]['count']
                        amount = data[employee][pattern]['amount']
                        value = f"{count}/{amount:.2f}"
                        
                        # 累计花型总计
                        pattern_totals[pattern]['count'] += count
                        pattern_totals[pattern]['amount'] += amount
                        employee_total += amount
                    else:
                        value = "0/0.00"
                    
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.alignment = center_alignment
                    cell.border = border
                
                # 写入员工合计
                cell = ws.cell(row=row_num, column=len(headers), value=f"{employee_total:.2f}")
                cell.alignment = center_alignment
                cell.border = border
                
                row_num += 1
            
            # 写入总计行
            ws.cell(row=row_num, column=1, value="合计").font = total_font
            ws.cell(row=row_num, column=1).fill = total_fill
            ws.cell(row=row_num, column=1).border = border
            
            grand_total = 0
            for col, pattern in enumerate(patterns, 2):
                total_count = pattern_totals[pattern]['count']
                total_amount = pattern_totals[pattern]['amount']
                value = f"{total_count}/{total_amount:.2f}"
                
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = center_alignment
                cell.border = border
                
                grand_total += total_amount
            
            # 写入总合计
            cell = ws.cell(row=row_num, column=len(headers), value=f"{grand_total:.2f}")
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # 调整列宽
            ws.column_dimensions['A'].width = 12
            for col in range(2, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 编码为Base64
            file_content = base64.b64encode(output.read()).decode('utf-8')
            
            # 生成文件名
            filename = f"工资汇总_{year}年{month:02d}月.xlsx"
            
            return {
                'success': True,
                'message': '导出成功',
                'data': {
                    'filename': filename,
                    'file_content': file_content,
                    'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'导出失败: {str(e)}'
            }
    
    def export_daily_summary(self, start_date: str = None, end_date: str = None) -> Dict[str, Any]:
        """
        导出日工资汇总Excel
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            导出结果
        """
        try:
            # 获取日汇总数据
            records = self.db.get_daily_summary(start_date, end_date)
            
            if not records:
                return {
                    'success': False,
                    'message': '没有找到相关记录'
                }
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "日工资汇总"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 创建表头
            headers = ['记录日期', '员工姓名', '花型名称', '针数', '工资金额']
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # 写入数据行
            row_num = 2
            for record in records:
                ws.cell(row=row_num, column=1, value=record['record_date']).border = border
                ws.cell(row=row_num, column=2, value=record['employee_name']).border = border
                ws.cell(row=row_num, column=3, value=record['pattern_name']).border = border
                
                cell = ws.cell(row=row_num, column=4, value=record['total_count'])
                cell.alignment = center_alignment
                cell.border = border
                
                cell = ws.cell(row=row_num, column=5, value=f"{record['total_amount']:.2f}")
                cell.alignment = center_alignment
                cell.border = border
                
                row_num += 1
            
            # 调整列宽
            column_widths = [12, 12, 12, 10, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 编码为Base64
            file_content = base64.b64encode(output.read()).decode('utf-8')
            
            # 生成文件名
            if start_date and end_date:
                filename = f"日工资汇总_{start_date}至{end_date}.xlsx"
            elif start_date:
                filename = f"日工资汇总_{start_date}开始.xlsx"
            else:
                filename = f"日工资汇总_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            return {
                'success': True,
                'message': '导出成功',
                'data': {
                    'filename': filename,
                    'file_content': file_content,
                    'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'导出失败: {str(e)}'
            }
    
    def create_import_template(self) -> Dict[str, Any]:
        """
        创建导入模板Excel文件
        
        Returns:
            模板文件
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "工资记录导入模板"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 创建表头
            headers = ['员工姓名', '花型名称', '记录日期', '针数', '特殊标记', '备注']
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # 写入示例数据
            example_data = [
                ['张三', '盖布', '2025-01-01', 1000, '', ''],
                ['李四', '浮花', '2025-01-01', 1500, '', '加班'],
                ['王五', '棉盖', '2025-01-02', 2000, '', '']
            ]
            
            for row_idx, row_data in enumerate(example_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx == 3:  # 日期列
                        cell.alignment = center_alignment
                    elif col_idx == 4:  # 针数列
                        cell.alignment = center_alignment
            
            # 调整列宽
            column_widths = [12, 12, 12, 10, 12, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 编码为Base64
            file_content = base64.b64encode(output.read()).decode('utf-8')
            
            return {
                'success': True,
                'message': '模板创建成功',
                'data': {
                    'filename': '工资记录导入模板.xlsx',
                    'file_content': file_content,
                    'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'模板创建失败: {str(e)}'
            }
